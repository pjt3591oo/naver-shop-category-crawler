import requests
from bs4 import BeautifulSoup
from time import sleep
from openpyxl import Workbook

wb = Workbook()
sheet1 = wb.active

CATEGORY_START = 50000000
CATEGORY_COUNT = 10

BASE_URL = 'https://search.shopping.naver.com'
CATEGORY_URL = "https://search.shopping.naver.com/category/category.nhn?cat_id=%s"

row = 1

for i in range(0, CATEGORY_COUNT+1):
  cat_id = CATEGORY_START+i
  url = CATEGORY_URL%(cat_id)
  category_page = requests.get(url)
  sleep(1)
  
  soup = BeautifulSoup(category_page.content, "lxml")
  top_text = soup.select('#content .category_tit')[0].text
  middle_categories = soup.select('.category_col')
  
  for middle_category in middle_categories:
    middle_text = middle_category.select('.category_cell h3 a strong')[0].text
    middle_link = middle_category.select('.category_cell h3 a')[0].get('href')
    bottom_categories = middle_category.select('ul.category_list > li:not(.tit_info)')

    if not len(bottom_categories) :
      print("%d, %s, %s, %s"%(cat_id, top_text, middle_text, middle_link))
      sheet1.cell(row=row, column=1).value = '%s'%(cat_id)
      sheet1.cell(row=row, column=2).value = '%s'%(top_text)
      sheet1.cell(row=row, column=3).value = '%s'%(middle_text)
      sheet1.cell(row=row, column=6).value = '%s%s'%(BASE_URL, middle_link)
      row += 1

    for bottom_category in bottom_categories:
      bottom_tab = bottom_category.select('a')[0]
      bottom_text = bottom_tab.text
      bottom_link = bottom_tab.get('href')
      leaf_categories = bottom_category.select('ul > li')
  
      for leaf_category in leaf_categories:
        leaf_tab = leaf_category('a')[0]
        leaf_href = leaf_tab.get('href')
        leaf_text = leaf_tab.text
        print("%d, %s, %s, %s %s %s"%(cat_id, top_text, middle_text, bottom_text, leaf_text, leaf_href))
        sheet1.cell(row=row, column=1).value = '%s'%(cat_id)
        sheet1.cell(row=row, column=2).value = '%s'%(top_text)
        sheet1.cell(row=row, column=3).value = '%s'%(middle_text)
        sheet1.cell(row=row, column=4).value = '%s'%(bottom_text)
        sheet1.cell(row=row, column=5).value = '%s'%(leaf_text)
        sheet1.cell(row=row, column=6).value = '%s%s'%(BASE_URL, middle_link)
        row += 1

wb.save(filename='naver_category.xlsx')